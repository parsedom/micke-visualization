import streamlit as st
import boto3
import pandas as pd
from datetime import datetime, timedelta
import json
import pytz
import plotly.express as px
from boto3.dynamodb.conditions import Key, Attr
import os
import hashlib
from decimal import Decimal
import hmac
from st_aggrid import AgGrid, GridOptionsBuilder
import io
import time
import uuid
import email as _email_mod
import email.mime.multipart
import email.mime.text
import email.mime.base
import email.encoders as _enc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.message import EmailMessage

st.set_page_config(
    page_title="Hotel Booking Dashboard",
    page_icon="🏨",
    layout="wide",
    initial_sidebar_state="expanded"
)

AVAILABLE_BOARDS = [
    "price_dashboard",
    "historical_calendar",
    "matrix_automation",
]
 
BOARD_LABELS = {
    "price_dashboard":    "📊 Price Dashboard",
    "historical_calendar":"📅 Historical Price Calendar",
    "matrix_automation":  "📤 Matrix Automation",
}

ALL_LOCATIONS = ["tampere", "oulu", "rauma", "turku", "jyvaskyla","vaasa","seinajoki"]

# ==================== CONFIGURATION & CONSTANTS ====================
ZONE1_HOTELS = [
    "Courtyard Tampere City",
    "Forenom Aparthotel Tampere Kauppakatu",
    "H28 - Hotel, Apartments and Suites by UHANDA",
    "Holiday Club Tampereen Kehräämö",
    "Holiday Inn Tampere - Central Station by IHG",
    "Hotel Kauppi",
    "Hotelli Vaakko - Hotel and Apartments by UHANDA",
    "Lapland Hotels Arena",
    "Lapland Hotels Tampere",
    "Lillan Hotel & Kök",
    "Original Sokos Hotel Ilves Tampere",
    "Original Sokos Hotel Villa Tampere",
    "Radisson Blu Grand Hotel Tammer",
    "Scandic Rosendahl",
    "Scandic Tampere City",
    "Scandic Tampere Hämeenpuisto",
    "Scandic Tampere Koskipuisto",
    "Scandic Tampere Station",
    "Solo Sokos Hotel Torni Tampere",
    "Unity Tampere - A Studio Hotel"
]

ZONE2_HOTELS = [
    "Courtyard Tampere City",
    "Forenom Aparthotel Tampere Kauppakatu",
    "H28 - Hotel, Apartments and Suites by UHANDA",
    "Holiday Club Tampereen Kehräämö",
    "Holiday Inn Tampere - Central Station by IHG",
    "Hotel Citi Inn",
    "Hotel Hermica",
    "Hotel Homeland",
    "Hotel Kauppi",
    "Hotelli Vaakko - Hotel and Apartments by UHANDA",
    "Hotelli Ville",
    "Lapland Hotels Arena",
    "Lapland Hotels Tampere",
    "Lillan Hotel & Kök",
    "Mango Hotel",
    "Omena Hotel Tampere",
    "Original Sokos Hotel Ilves Tampere",
    "Original Sokos Hotel Villa Tampere",
    "Radisson Blu Grand Hotel Tammer",
    "Scandic Eden Nokia",
    "Scandic Rosendahl",
    "Scandic Tampere City",
    "Scandic Tampere Hämeenpuisto",
    "Scandic Tampere Koskipuisto",
    "Scandic Tampere Station",
    "Solo Sokos Hotel Torni Tampere",
    "Unity Tampere - A Studio Hotel",
    "Uumen Hotels - Tampere, Finlayson"
]

ZONE3_HOTELS = [
    "H28 - Hotel, Apartments and Suites by UHANDA",
    "Forenom Aparthotel Tampere Kauppakatu",
    "Forenom Serviced Apartments Tampere Pyynikki",
    "Scandic Tampere Hämeenpuisto",
    "Uumen Hotels - Tampere, Finlayson",
    "Scandic Tampere Koskipuisto",
    "Radisson Blu Grand Hotel Tammer",
    "Original Sokos Hotel Ilves Tampere",
    "Omena Hotel Tampere",
    "Scandic Tampere City",
    "Holiday Inn Tampere - Central Station by IHG",
    "Solo Sokos Hotel Torni Tampere",
    "Scandic Tampere Station",
    "Lapland Hotels Arena",
    "Hotel Citi Inn",
    "Scandic Rosendahl",
    "Original Sokos Hotel Villa Tampere",
    "Lapland Hotels Tampere",
    "Unity Tampere - A Studio Hotel",
    "Courtyard Tampere City",
    "Hotelli Ville",
    "Hotel Homeland",
    "Mango Hotel",
    "Holiday Club Tampereen Kehräämö",
    "Hotel Kauppi",
    "Varala Sports & Nature Hotel",
    "Lillan Hotel & Kök",
    "Forenom Aparthotel Tampere Kaleva",
    "Norlandia Tampere Hotel",
    "Hotelli Vaakko - Hotel and Apartments by UHANDA",
    "Hotel Lamminpää",
    "Hotel Hermica",
    "Oma Hotelli",
    "Scandic Eden Nokia",
    "Hotelli Iisoppi",
    "Hotelli Kuohu Kangasala - Hotel and Apartments by UHANDA",
    "Hotel Urkin Pillopirtti",
    "Hotel Waltikka",
    "Aapiskukko Hotel",
    "Hotel Ackas",
    "Hotelli Sointula"
]

Alert_Comparison = [
    "Unity Tampere - A Studio Hotel",
    "Scandic Tampere Station",
    "Scandic Tampere Koskipuisto",
    "Scandic Tampere Hämeenpuisto",
    "Scandic Tampere City",
    "Scandic Rosendahl",
    "Original Sokos Hotel Villa Tampere",
    "Hotelli Vaakko - Hotel and Apartments by UHANDA",
    "Hotel Kauppi",
    "Holiday Inn Tampere - Central Station by IHG",
    "Holiday Club Tampereen Kehräämö",
    "H28 - Hotel, Apartments and Suites by UHANDA",
    "Courtyard Tampere City"
]

aws_key = st.secrets["AWS_ACCESS_KEY_ID"]
aws_secret = st.secrets["AWS_SECRET_ACCESS_KEY"]
region = st.secrets["AWS_DEFAULT_REGION"]

# STD_TOP_VALUE = 600


dynamodb = boto3.resource(
    'dynamodb',
    aws_access_key_id=aws_key,
    aws_secret_access_key=aws_secret,
    region_name=region
)

table = dynamodb.Table('HotelPrices')
table_calender = dynamodb.Table('HotelPricesCalendar')
table_user = dynamodb.Table('MickeUser')
table_color = dynamodb.Table('micke_color_config')
table_logs = dynamodb.Table('MickeLoginLogs')
table_config = dynamodb.Table('MickeAppConfig')
table_zones = dynamodb.Table('MickeZones')
table_emails = dynamodb.Table('MickeEmailList') 

# ==================== ZONE HELPER FUNCTIONS ====================

@st.cache_data(ttl=60)
def get_zones_for_location(location: str) -> list:
    """
    Return all zones for a given location sorted by sort_order → zone_name.
    PK format: "zone_name#location" — we scan and filter by the location attribute.
    Cached 60 s to avoid hammering DynamoDB on every widget interaction.
    """
    try:
        response = table_zones.scan(FilterExpression=Attr('location').eq(location))
        items = response.get('Items', [])
        while 'LastEvaluatedKey' in response:
            response = table_zones.scan(
                FilterExpression=Attr('location').eq(location),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response.get('Items', []))
    except Exception as e:
        st.warning(f"Could not load zones for {location}: {e}")
        return []

    items.sort(key=lambda x: (int(x.get('sort_order', 9999)), x.get('zone_name', '')))
    return items


@st.cache_data(ttl=60)
def get_all_zones() -> list:
    """Return every zone row sorted by location → sort_order → zone_name."""
    try:
        response = table_zones.scan()
        items = response.get('Items', [])
        while 'LastEvaluatedKey' in response:
            response = table_zones.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
            items.extend(response.get('Items', []))
        items.sort(key=lambda x: (x.get('location', ''),
                                   int(x.get('sort_order', 9999)),
                                   x.get('zone_name', '')))
        return items
    except Exception as e:
        st.warning(f"Could not load zones: {e}")
        return []


# def _make_zone_pk(zone_name: str, location: str) -> str:
#     """Build the partition key string: 'zone_name#location'."""
#     return f"{zone_name}#{location}"


def save_zone(zone_name: str, location: str, hotels: list, sort_order: int = 0) -> bool:
    """
    Upsert a zone. PK = "zone_name#location" (single partition key, no sort key).
    Storing zone_name and location as separate attributes too so we can filter/display them.
    Clears cache after saving so the UI reflects the change immediately.
    """
    try:
        table_zones.put_item(Item={
            'zone_name#location': f"{zone_name}#{location}",
            'zone_name':  zone_name,
            'location':   location,
            'hotels':     hotels,
            'updated_at': datetime.now().isoformat()
        })
        get_zones_for_location.clear()
        get_all_zones.clear()
        return True
    except Exception as e:
        st.error(f"Failed to save zone: {e}")
        return False


def delete_zone(zone_name: str, location: str) -> bool:
    """Delete a zone using its PK: 'zone_name#location'."""
    try:
        table_zones.delete_item(Key={'zone_name#location': f"{zone_name}#{location}"})
        get_zones_for_location.clear()
        get_all_zones.clear()
        return True
    except Exception as e:
        st.error(f"Failed to delete zone: {e}")
        return False


def _resolve_zone_hotels(zone_name: str, location: str) -> list:
    """
    Fetch hotel list for a zone from DynamoDB using PK 'zone_name#location'.
    Falls back to legacy hardcoded lists for backward-compat
    (existing sessions that stored "zone1"/"zone2" etc. still work).
    """
    try:
        resp = table_zones.get_item(Key={'zone_name#location': f"{zone_name}#{location}"})
        item = resp.get('Item')
        if item:
            return item.get('hotels', [])
    except Exception:
        pass
    # Legacy fallback
    legacy = {
        'zone1': ZONE1_HOTELS, 'zone2': ZONE2_HOTELS,
        'zone3': ZONE3_HOTELS, 'alert': Alert_Comparison
    }
    return legacy.get(zone_name, ZONE1_HOTELS)

## ==================== EMAIL CONFIGURATION FUNCTIONS ====================

@st.cache_data(ttl=30)
def get_saved_emails() -> list:
    """Return all saved email addresses from the global list, sorted."""
    try:
        resp  = table_emails.scan()
        items = resp.get('Items', [])
        while 'LastEvaluatedKey' in resp:
            resp   = table_emails.scan(ExclusiveStartKey=resp['LastEvaluatedKey'])
            items.extend(resp.get('Items', []))
        return sorted([i['email'] for i in items if i.get('email')])
    except Exception as e:
        st.warning(f"Could not load email list: {e}")
        return []
 
 
def save_email(email_address: str) -> bool:
    try:
        table_emails.put_item(Item={
            'email':      email_address.strip().lower(),
            'added_at':   datetime.now().isoformat(),
            'added_by':   st.session_state.get('authenticated_user', 'admin'),
        })
        get_saved_emails.clear()
        return True
    except Exception as e:
        st.error(f"Failed to save email: {e}")
        return False
 
 
def delete_email(email_address: str) -> bool:
    try:
        table_emails.delete_item(Key={'email': email_address})
        get_saved_emails.clear()
        return True
    except Exception as e:
        st.error(f"Failed to delete email: {e}")
        return False
    
# ==================== COLOR CONFIGURATION FUNCTIONS ====================

def get_color_presets_for_location(location):
    """Get all color config names that apply to a specific location."""
    try:
        response = table_color.scan(
            FilterExpression=Attr('locations').contains(location)
        )
        
        items = response.get('Items', [])
        
        while 'LastEvaluatedKey' in response:
            response = table_color.scan(
                FilterExpression=Attr('locations').contains(location),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response.get('Items', []))
        
        config_names = [item.get('color_config_name') for item in items if item.get('color_config_name')]
        return sorted(list(set(config_names)))
    
    except Exception as e:
        st.warning(f"Error loading presets for {location}: {e}")
        return []


def get_color_config_by_name(color_config_name):
    """Get color ranges using the unique config name."""
    try:
        response = table_color.query(
            KeyConditionExpression=Key('color_config_name').eq(color_config_name)
        )
        
        if response.get('Items'):
            item = response['Items'][0]
            ranges = item.get('ranges', [])
            
            converted_ranges = []
            for r in ranges:
                converted_ranges.append({
                    'min': float(r['min_value']),
                    'max': float(r['max_value']),
                    'color': r['color']
                })
            
            return converted_ranges
        else:
            return get_default_color_ranges().get('zone1', [])
    
    except Exception as e:
        st.error(f"Error loading color config '{color_config_name}': {e}")
        return get_default_color_ranges().get('zone1', [])

def get_std_top_value():
    try:
        response = table_config.get_item(Key={"config_key": "std_top_value"})
        item = response.get("Item")
        if item:
            return int(item.get("value", 600))
    except Exception:
        pass
    return 600

def save_std_top_value(value: int):
    try:
        table_config.put_item(Item={"config_key": "std_top_value", "value": value})
        return True
    except Exception as e:
        st.error(f"Failed to save config: {e}")
        return False

    
def check_password():
    """Returns True if the user has the correct password using DynamoDB."""

    def password_entered():
        username = st.session_state.get("username")
        password = st.session_state.get("password")

        if not username or not password:
            st.session_state["password_correct"] = False
            return

        try:
            response = table_user.get_item(
                Key={"username": username}
            )
            user = response.get("Item")

            if user and hmac.compare_digest(password, user.get("password", "")):
                st.session_state["password_correct"] = True
                st.session_state["authenticated_user"] = user["username"]
                st.session_state["access"] = user.get("access", "user")
                if user.get("access") == "admin":
                    st.session_state["boards"] = AVAILABLE_BOARDS
                    st.session_state["locations"] = ["tampere", "oulu", "rauma", "turku", "jyvaskyla","vaasa","seinajoki"]
                else:
                    st.session_state["boards"] = user.get("boards", [])
                    st.session_state["locations"] = user.get("locations", [])

                helsinki_tz = pytz.timezone('Europe/Helsinki')
                finland_now = datetime.now(helsinki_tz)

                table_user.update_item(
                    Key={"username": user["username"]},
                    UpdateExpression="SET last_login = :l",
                    ExpressionAttributeValues={
                        ":l": finland_now.strftime("%d/%m/%Y")
                    }
                )

                try:
                    table_logs.put_item(
                        Item={
                            "username": user["username"],
                            "login_ts": finland_now.isoformat(),
                            "login_date": finland_now.strftime("%Y-%m-%d")
                        }
                    )
                except Exception as e:
                    st.warning("Login logged with warning")
                    print(e)

                del st.session_state["password"]
                del st.session_state["username"]
            else:
                st.session_state["password_correct"] = False

        except Exception as e:
            st.session_state["password_correct"] = False
            st.error(f"Login error: {e}")

    if st.session_state.get("password_correct", False):
        return True

    st.markdown("""
    <div style="max-width: 400px; margin: 5rem auto; padding: 2rem; 
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                border-radius: 15px;">
        <h1 style="color: white; text-align: center;">🏨 Hotel Dashboard Login</h1>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.text_input("Username", key="username")
        st.text_input("Password", type="password", key="password")
        st.button("Login", on_click=password_entered, type="primary", use_container_width=True)

        if "password_correct" in st.session_state and not st.session_state["password_correct"]:
            st.error("😞 User not known or password incorrect")

    return False

def logout():
    """Clear session state for logout"""
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

if not check_password():
    st.stop()


# ==================== STYLING ====================
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .stDataFrame {
        border: 1px solid #e6e6e6;
        border-radius: 8px;
    }
    .hotel-selector {
        background: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .user-info {
        background: #e8f4fd;
        padding: 1rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        border-left: 4px solid #1f77b4;
    }
    .calendar-table {
        width: 100%;
        border-collapse: collapse;
        margin: 20px 0;
        border: 2px solid #000;
    }
    .calendar-table td {
        border: 1px solid #000;
        padding: 0;
        height: 80px;
        position: relative;
    }
    .week-label {
        text-align: center;
        font-weight: bold;
        padding: 10px;
        background-color: #f5f5f5;
        border-right: 2px solid #000;
        width: 150px !important;
        height: 80px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 12px;
        color: #000;
    }
    .day-label {
        text-align: center;
        font-weight: bold;
        padding: 8px;
        background-color: #f5f5f5;
        border: 1px solid #000;
        min-height: 30px;
        flex: 1;
        color: #000;
    }
    .cell-content {
        width: 100%;
        height: 100%;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        color: white;
        font-weight: bold;
        text-align: center;
        padding: 10px;
        box-sizing: border-box;
        color: #000;
    }
    .cell-date {
        font-size: 12px;
        margin-bottom: 5px;
    }
    .cell-value {
        font-size: 18px;
    }
   .empty-cell {
        background-color: #ffffff;
    }
    /* Prevent multiselect from expanding vertically */
    div[data-baseweb="select"] > div {
        max-height: 38px !important;
        overflow-y: auto !important;
    }
    /* Keep the dropdown list normal */
    div[data-baseweb="popover"] div[data-baseweb="select"] > div {
        max-height: none !important;
    }
</style>
""", unsafe_allow_html=True)

def get_default_color_ranges():
    """Get default color ranges for all zones"""
    return {
        "zone1": [
            {'min': 0.0, 'max': 124.99, 'color': '#08306b'},
            {'min': 125.0, 'max': 134.99, 'color': '#2171b5'},
            {'min': 135.0, 'max': 144.99, 'color': '#a2cff8'},
            {'min': 145.0, 'max': 154.99, 'color': '#ffffff'},
            {'min': 155.0, 'max': 164.99, 'color': '#ffa0a0'},
            {'min': 165.0, 'max': 199.99, 'color': '#f86868'},
            {'min': 200.0, 'max': 249.99, 'color': '#d81919'},
            {'min': 250.0, 'max': 999999.0, 'color': '#000000'}
        ]
    }


def get_text_color_from_background(hex_color):
    """
    Determine if text should be white or black based on background color brightness.
    Uses relative luminance calculation (WCAG standard).
    """
    hex_color = hex_color.lstrip('#')
    
    r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    def adjust_color(c):
        c = c / 255.0
        if c <= 0.03928:
            return c / 12.92
        else:
            return ((c + 0.055) / 1.055) ** 2.4
    
    r_adjusted = adjust_color(r)
    g_adjusted = adjust_color(g)
    b_adjusted = adjust_color(b)
    
    luminance = 0.2126 * r_adjusted + 0.7152 * g_adjusted + 0.0722 * b_adjusted
    
    return "#000000" if luminance > 0.5 else "#FFFFFF"


def get_color_from_price_ranges(value, price_ranges):
    """Generate color based on price value and defined ranges."""
    if not price_ranges:
        return "rgb(150, 150, 150)"
    
    sorted_ranges = sorted(price_ranges, key=lambda x: x['min'])
    
    for range_item in sorted_ranges:
        if range_item['min'] <= value <= range_item['max']:
            return range_item['color']
    
    if value < sorted_ranges[0]['min']:
        return sorted_ranges[0]['color']
    else:
        return sorted_ranges[-1]['color']


def make_x_label(price_date_series):
    """Generate stacked x-axis labels: DAY / DD / Mon, red for weekends"""
    day_abbr = price_date_series.dt.strftime('%a').str.upper()
    day_num = price_date_series.dt.strftime('%d').str.lstrip('0')
    month_abbr = price_date_series.dt.strftime('%b')
    dayofweek = price_date_series.dt.dayofweek
    is_weekend = (dayofweek == 4) | (dayofweek == 5)
    weekend_color = "#ff6b6b"
    labels = []
    for i in range(len(price_date_series)):
        if is_weekend.iloc[i]:
            label = (
                f'<b><span style="color:{weekend_color}">' + day_abbr.iloc[i] + '</span></b><br>' +
                f'<b><span style="color:{weekend_color}">' + day_num.iloc[i] + '</span></b><br>' +
                f'<b><span style="color:{weekend_color}">' + month_abbr.iloc[i] + '</span></b>'
            )
        else:
            label = (
                '<b>' + day_abbr.iloc[i] + '</b><br>' +
                '<b>' + day_num.iloc[i] + '</b><br>' +
                '<b>' + month_abbr.iloc[i] + '</b>'
            )
        labels.append(label)

    return pd.Series(labels, index=price_date_series.index)

    
# ==================== QUERY FUNCTIONS ====================
def query_hotels(filters, date_range, scraped_date_start, scraped_date_end):
    """Query DynamoDB for hotel prices based on filters and date ranges."""
    location = filters.get('location')
    time = filters.get('time')
    persons = filters.get('persons')
    nights = filters.get('nights')
    
    if date_range:
        dates = date_range.split(' - ')
        def convert_date_format(date_str):
            day, month, year = date_str.split('-')
            return f"{year}-{month}-{day}"
        
        checkin_start = convert_date_format(dates[0])
        checkin_end = convert_date_format(dates[1])
    else:
        checkin_start = None
        checkin_end = None

    partition_key = f"{location}#{persons}#{nights}#{time}"
    
    filter_expression = None
    if checkin_start and checkin_end:
        filter_expression = Attr('checkin_date').between(checkin_start, checkin_end)
    
    all_items = []
    
    try:
        key_condition = (
            Key('location#persons#nights#time').eq(partition_key) &
            Key('scraped_date#hotel_id#checkin_date#checkout_date')
                .between(f"{scraped_date_start}#", f"{scraped_date_end}~")
        )
        
        response = table.query(
            KeyConditionExpression=key_condition,
            FilterExpression=filter_expression
        )
        
        items = response['Items']
        
        while 'LastEvaluatedKey' in response:
            response = table.query(
                KeyConditionExpression=key_condition,
                FilterExpression=filter_expression,
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response['Items'])
        
        all_items.extend(items)
        
        transformed_items = []
        for item in all_items:
            transformed_items.append({
                'name': item.get('hotel_name', ''),
                'price': item.get('price', 0),
                'price_date': item.get('checkin_date', ''),
                'scrape_date': item.get('scraped_date', ''),
                'location': item.get('location', ''),
                'persons': item.get('persons', 0),
                'nights': item.get('nights', 0),
                'time': item.get('time', ''),
                'review_score': item.get('review_score', 0),
                'city': item.get('city', ''),
                'distance': item.get('distance', ''),
                'hotel_url': item.get('hotel_url', ''),
                'breakfast_included': item.get('breakfast_included', False),
                'free_cancellation': item.get('free_cancellation', False)
            })
        
        return transformed_items
    
    except Exception as e:
        st.error(f"Error querying DynamoDB: {str(e)}")
        return []

# def query_calendar_data(price_start_date, price_end_date, zone_filter="zone1", location="tampere"):
#     """Query data for calendar heatmap."""
#     price_dates = []
#     d = price_start_date
#     while d <= price_end_date:
#         price_dates.append(d.strftime("%Y-%m-%d"))
#         d += timedelta(days=1)
    
#     metrics = {
#         'availability': {},
#         'price_avg': {},
#         'free_cancel_avg': {}
#     }

#     for idx, pdate in enumerate(price_dates, 1):
#         partition_key = f"{location}#{zone_filter}#{pdate}"
#         try:
#             response = table_calender.get_item(
#                 Key={'location#zone#checkin_date': partition_key}
#             )
#             item = response.get('Item')
#             if item:
#                 metrics["availability"][pdate] = item.get("availability")
#                 metrics["free_cancel_avg"][pdate] = item.get("free_cancel_avg")
#                 metrics["price_avg"][pdate] = item.get("price_avg")
#         except Exception as e:
#             st.error(f"Error querying DynamoDB: {str(e)}")
#             return metrics

#     return metrics

def query_calendar_hotels(date_range, scraped_date_start, scraped_date_end):
    location = "tampere"
    time = "morning"
    persons = 2
    nights = 1

    if date_range:
        dates = date_range.split(' - ')
        day, month, year = dates[0].split('-')
        checkin_date = f"{year}-{month}-{day}"
    else:
        return []

    partition_key = f"{location}#{persons}#{nights}#{time}"

    all_items = []

    try:
        key_condition = (
            Key('location#persons#nights#time').eq(partition_key) &
            Key('checkin_date#scraped_date').between(
                f"{checkin_date}#{scraped_date_start}",
                f"{checkin_date}#{scraped_date_end}~"
            )
        )

        response = table.query(
            IndexName='hotel_prices_by_checkin_scraped',
            KeyConditionExpression=key_condition
        )

        items = response['Items']

        while 'LastEvaluatedKey' in response:
            response = table.query(
                IndexName='hotel_prices_by_checkin_scraped',
                KeyConditionExpression=key_condition,
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response['Items'])

        all_items.extend(items)

        transformed_items = []
        for item in all_items:
            transformed_items.append({
                'name': item.get('hotel_name', ''),
                'price': item.get('price', 0),
                'price_date': item.get('checkin_date', ''),
                'scrape_date': item.get('scraped_date', ''),
                'location': item.get('location', ''),
                'persons': item.get('persons', 0),
                'nights': item.get('nights', 0),
                'time': item.get('time', ''),
                'review_score': item.get('review_score', 0),
                'city': item.get('city', ''),
                'distance': item.get('distance', ''),
                'hotel_url': item.get('hotel_url', ''),
                'breakfast_included': item.get('breakfast_included', False),
                'free_cancellation': item.get('free_cancellation', False)
            })

        return transformed_items

    except Exception:
        return []

def query_calendar_data(price_start_date, price_end_date, zone_filter="zone1", location="tampere"):
    """Query data for calendar heatmap."""
    price_dates = []
    d = price_start_date
    while d <= price_end_date:
        price_dates.append(d.strftime("%Y-%m-%d"))  # YYYY-MM-DD (DB format)
        d += timedelta(days=1)
    
    time = "morning"
    persons = 2
    nights = 1
    partition_key = f"{location}#{persons}#{nights}#{time}"
    
    metrics = {
        'availability': {},
        'price_avg': {},
        'free_cancel_avg': {}
    }
    
    zone_mapping = {
        'zone1': ZONE1_HOTELS,
        'zone2': ZONE2_HOTELS,
        'zone3': ZONE3_HOTELS,
        'alert': Alert_Comparison
    }
    
    selected_zone = zone_mapping.get(zone_filter, ZONE1_HOTELS)
    
    for idx, pdate in enumerate(price_dates, 1):
        # For this price date, scrape window is 30 days before it
        price_dt = datetime.strptime(pdate, "%Y-%m-%d")
        scrape_start_dt = price_dt - timedelta(days=30)
        
        scraped_start = scrape_start_dt.strftime("%Y-%m-%d")
        scraped_end = pdate

        price_ddmmyyyy = price_dt.strftime("%d-%m-%Y")
        date_range_for_query = f"{price_ddmmyyyy} - {price_ddmmyyyy}"


        results = query_calendar_hotels(
            date_range=date_range_for_query,
            scraped_date_start=scraped_start,
            scraped_date_end=scraped_end
        )

        # Query for price averages (30-day scrape window)
        if not results:
            metrics['free_cancel_avg'][pdate] = 0
            metrics['price_avg'][pdate] = 0
            metrics['availability'][pdate] = 0
            continue

        df = pd.DataFrame(results)
        df['price'] = pd.to_numeric(df['price'], errors='coerce')
        df = df.dropna(subset=['price'])
        fc_df = df.copy()
        wo_df = df[(df['breakfast_included'] == False) & (df['free_cancellation'] == False)]

        fc_df = df[(df['breakfast_included'] == False) & (df['free_cancellation'] == True)]

        wo_unique_hotels = sorted(wo_df['name'].unique())
        fc_unique_hotels = sorted(fc_df['name'].unique())

        wo_available_zone1 = [hotel for hotel in selected_zone if hotel in wo_unique_hotels]
        fc_available_zone1 = [hotel for hotel in selected_zone if hotel in fc_unique_hotels]

        wo_df_zone1 = wo_df[wo_df['name'].isin(wo_available_zone1)]
        fc_df_zone1 = fc_df[fc_df['name'].isin(fc_available_zone1)]

        wo_avg = round(wo_df_zone1['price'].mean(),2) if not wo_df_zone1.empty else 0
        fc_avg = round(fc_df_zone1['price'].mean(),2) if not fc_df_zone1.empty else 0

        total_records = len(df)
        price_ddmmyyyy_avail = price_dt.strftime("%d-%m-%Y")
        date_range_for_avail = f"{price_ddmmyyyy_avail} - {price_ddmmyyyy_avail}"
        
        results_avail = query_calendar_hotels(
            date_range=date_range_for_avail,
            scraped_date_start=pdate,
            scraped_date_end=pdate
        )
        
        if results_avail:
            df_avail = pd.DataFrame(results_avail)
            df_avail['price'] = pd.to_numeric(df_avail['price'], errors='coerce')
            df_avail = df_avail.dropna(subset=['price'])
            df_avail = df_avail[(df_avail['breakfast_included'] == False)]
            unique_hotels_avail = sorted(df_avail['name'].unique())
            available_zone1 = [hotel for hotel in selected_zone if hotel in unique_hotels_avail]
        else:
            available_zone1 = []
        
        num_zone1_available = len(available_zone1)

        TOTAL_ZONE1 = len(selected_zone)
        zone1_avail_pct = round((num_zone1_available / TOTAL_ZONE1) * 100, 1) if TOTAL_ZONE1 > 0 else 0

        metrics['free_cancel_avg'][pdate] = fc_avg
        metrics['price_avg'][pdate] = wo_avg
        metrics['availability'][pdate] = zone1_avail_pct
    
    return metrics

def _query_matrix_data(location: str, persons: int, time_val: str,
                        start_date, days_forward: int) -> list:
    """
    Query HotelPrices for a single persons/time combo.
    - scrape_date = start_date (same day the user picks)
    - checkin window = start_date → start_date + days_forward - 1
    - nights = 1
    - No breakfast/cancellation filter — returns all rows
    """
    scraped_date_str = start_date.strftime("%Y-%m-%d")

    end_date      = start_date + timedelta(days=days_forward - 1)
    checkin_start = start_date.strftime("%Y-%m-%d")
    checkin_end   = end_date.strftime("%Y-%m-%d")

    pk = f"{location}#{persons}#1#{time_val}"

    key_cond    = (
        Key("location#persons#nights#time").eq(pk) &
        Key("scraped_date#hotel_id#checkin_date#checkout_date")
            .between(f"{scraped_date_str}#", f"{scraped_date_str}~")
    )
    filter_expr = Attr("checkin_date").between(checkin_start, checkin_end)

    items = []
    try:
        resp = table.query(KeyConditionExpression=key_cond,
                           FilterExpression=filter_expr)
        items.extend(resp["Items"])
        while "LastEvaluatedKey" in resp:
            resp = table.query(
                KeyConditionExpression=key_cond,
                FilterExpression=filter_expr,
                ExclusiveStartKey=resp["LastEvaluatedKey"]
            )
            items.extend(resp["Items"])
    except Exception as e:
        st.error(f"DynamoDB query error: {e}")

    return [
        {
            "name":               item.get("hotel_name", ""),
            "price":              item.get("price", 0),
            "checkin_date":       item.get("checkin_date", ""),
            "hotel_url":          item.get("hotel_url", ""),
            "review_score":       item.get("review_score", ""),
            "city":               item.get("city", ""),
            "distance":           item.get("distance", ""),
            "breakfast_included": item.get("breakfast_included", False),
            "free_cancellation":  item.get("free_cancellation", False),
            "persons":            persons,
        }
        for item in items
    ]
 
def _build_excel_matrix(df: pd.DataFrame, zone_name: str, location: str,
                         persons_list: list, single_sheet: str = None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font   = Font(bold=True, color="000000", name="Arial", size=9)
    hotel_font = Font(bold=True, name="Arial", size=9)
    data_font  = Font(name="Arial", size=9)
    url_font   = Font(name="Arial", size=9, color="0563C1", underline="single")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    META_COLS = ["URL", "Name", "Review Score", "City", "Distance",
                 "Breakfast Included", "Free Cancellation"]
    n_meta = len(META_COLS)

    df = df.copy()
    df["row_id"] = (
        df["name"].astype(str) + " | " +
        df["breakfast_included"].astype(str) + " | " +
        df["free_cancellation"].astype(str)
    )

    all_dates  = sorted(df["checkin_date"].unique())
    col_labels = [datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
                  for d in all_dates]
    all_headers = META_COLS + col_labels
    n_total     = len(all_headers)

    FILTERS = [
        ("No Extras",          False, False),
        ("Free Cancellation",  False, True),
        ("Breakfast Included", True,  False),
        ("Breakfast + Cancel", True,  True),
    ]

    def _write_sheet(ws, filtered_df):
        for ci, label in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = brd
        ws.row_dimensions[1].height = 28

        for ci, w in enumerate([45, 38, 12, 14, 22, 18, 18], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ci in range(n_meta + 1, n_total + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 12

        if filtered_df.empty:
            ws.cell(row=2, column=1, value="No data for this filter combination.").font = data_font
            return

        current_row = 2
        for persons in sorted(persons_list):
            p_df = filtered_df[filtered_df["persons"] == persons].copy()
            if p_df.empty:
                continue

            ordered = sorted(p_df["row_id"].unique())
            meta_lookup = (
                p_df.drop_duplicates(subset=["row_id"])
                    .set_index("row_id")[["hotel_url", "review_score", "city",
                                          "distance", "breakfast_included",
                                          "free_cancellation"]]
                    .to_dict("index")
            )
            pivot = p_df.pivot_table(
                index="row_id", columns="checkin_date",
                values="price", aggfunc="first"
            )

            for row_id in ordered:
                meta  = meta_lookup.get(row_id, {})
                hotel = row_id.split(" | ")[0]

                url = meta.get("hotel_url", "")
                uc  = ws.cell(row=current_row, column=1, value=url)
                if url:
                    uc.hyperlink = url
                    uc.font = url_font
                else:
                    uc.font = data_font
                uc.alignment = left
                uc.border    = brd

                nc = ws.cell(row=current_row, column=2, value=hotel)
                nc.font = hotel_font; nc.alignment = left; nc.border = brd

                rv = meta.get("review_score", "")
                rc = ws.cell(row=current_row, column=3, value=float(rv) if rv else "")
                rc.font = data_font; rc.alignment = center; rc.border = brd

                cc = ws.cell(row=current_row, column=4, value=meta.get("city", ""))
                cc.font = data_font; cc.alignment = left; cc.border = brd

                dc = ws.cell(row=current_row, column=5, value=meta.get("distance", ""))
                dc.font = data_font; dc.alignment = left; dc.border = brd

                bc = ws.cell(row=current_row, column=6,
                             value=str(meta.get("breakfast_included", False)))
                bc.font = data_font; bc.alignment = center; bc.border = brd

                fc = ws.cell(row=current_row, column=7,
                             value=str(meta.get("free_cancellation", False)))
                fc.font = data_font; fc.alignment = center; fc.border = brd

                for col_offset, date_str in enumerate(all_dates):
                    col_num = n_meta + col_offset + 1
                    try:
                        val = pivot.loc[row_id, date_str]
                        if pd.isna(val):
                            raise KeyError
                        pc = ws.cell(row=current_row, column=col_num,
                                     value=round(float(val), 1))
                        pc.number_format = "0.0"
                    except (KeyError, TypeError):
                        pc = ws.cell(row=current_row, column=col_num, value="")
                    pc.font = data_font; pc.alignment = center; pc.border = brd

                current_row += 1

        ws.freeze_panes = f"{get_column_letter(n_meta + 1)}2"

    if single_sheet:
        # Filtered mode one sheet with exactly the pre-filtered data
        ws = wb.create_sheet(title=single_sheet)
        _write_sheet(ws, df)
    else:
        # All data mode — 4 sheets split by breakfast/cancellation combo
        for sheet_label, bf, fc_flag in FILTERS:
            mask = (
                (df["breakfast_included"] == bf) &
                (df["free_cancellation"]  == fc_flag)
            )
            ws = wb.create_sheet(title=sheet_label)
            _write_sheet(ws, df[mask].copy())

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
 
 
def _send_matrix_email(recipients: list, subject: str, body: str,
                        xlsx_bytes: bytes, filename: str):
    """Send Excel matrix as attachment via Gmail SMTP."""
    

    sender_email   = st.secrets["GMAIL_SENDER"]   
    sender_password = st.secrets["GMAIL_APP_PASSWORD"] 

    msg            = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = sender_email
    msg["To"]      = ", ".join(recipients)
    msg.set_content(body)

    msg.add_attachment(
        xlsx_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)    


def get_color_from_availability(value, min_val, max_val):
    """Generate color based on value (green for high, red for low)."""
    if max_val == min_val:
        normalized = 0.5
    else:
        normalized = (value - min_val) / (max_val - min_val)
    
    if normalized < 0.33:
        r, g, b = 255, 100, 100  # Red
    elif normalized < 0.67:
        r, g, b = 255, 200, 100  # Orange
    else:
        r, g, b = 100, 200, 100  # Green
    
    return f"rgb({int(r)}, {int(g)}, {int(b)})"

# ==================== PAGE NAVIGATION ====================
boards   = st.session_state.get("boards", [])
is_admin = st.session_state.get("access") == "admin"
 
tab_labels = []
for b in AVAILABLE_BOARDS:
    if b == "matrix_automation":
        if is_admin:                              # admin-only tab
            tab_labels.append(BOARD_LABELS[b])
    elif b in boards:
        tab_labels.append(BOARD_LABELS[b])
 
if is_admin:
    tab_labels.append("🛠️ Admin Panel")
 
created_tabs = st.tabs(tab_labels)
 
tab1 = tab2 = tab_matrix = admin_panel = None
tab_index = 0
 
if "price_dashboard" in boards:
    tab1 = created_tabs[tab_index]; tab_index += 1
 
if "historical_calendar" in boards:
    tab2 = created_tabs[tab_index]; tab_index += 1
 
if is_admin:
    tab_matrix  = created_tabs[tab_index]; tab_index += 1
    admin_panel = created_tabs[tab_index]


# ==================== TAB 1: PRICE DASHBOARD ====================
if tab1:
    with tab1:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            <div class="main-header">
                <h1>🏨 Hotel Booking Price Dashboard</h1>
                <p>Analyze hotel prices across different dates and locations</p>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            st.markdown(f"""
            <div class="user-info">
                <p><strong>👤 Logged in as:</strong><br>{st.session_state.get('authenticated_user', 'Unknown')}</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("🚪 Logout", type="secondary", use_container_width=True):
                logout()

        with st.sidebar:
            st.markdown("### 🔧 Configuration")
            
            with st.expander("📍 Location & Booking Details", expanded=True):
                allowed_locations = st.session_state.get("locations", [])
                location = st.selectbox(
                    "Location",
                    allowed_locations,
                    index=0 if allowed_locations else None
                )
                persons = st.selectbox("Persons", [1, 2])
                nights = st.selectbox("Nights", [1,2,3, 7])
                time_of_day = st.selectbox("Time", ["morning", "evening"])
                breakfast_filter = st.checkbox("🍳 Include Breakfast Only", value=False)
                cancellation_filter = st.checkbox("✅ Free Cancellation", value=False)
            
            with st.expander("📅 Date Ranges", expanded=True):
                import datetime as _dt

                def year_has_week53(year):
                    try:
                        _dt.date.fromisocalendar(year, 53, 1)
                        return True
                    except ValueError:
                        return False

                current_year = datetime.now().year

                # ── View Dates (Scraped) ──────────────────────────────────────
                st.markdown("**View Dates**")
                view_mode = st.radio("View date input mode", ["📅 Calendar", "📆 Week"],
                                    horizontal=True, key="view_date_mode",
                                    label_visibility="collapsed")

                selected_view_weeks = []  # always initialize

                if view_mode == "📅 Calendar":
                    scraped_date_range = st.date_input(
                        "View Dates", value=[],
                        key="scrape_dates_unique",
                        label_visibility="collapsed"
                    )
                    scraped_week_selection = None
                else:
                    view_year_options = list(range(current_year - 2, current_year + 1))
                    view_year = st.selectbox(
                        "Year", view_year_options,
                        index=len(view_year_options) - 1,
                        key="view_year_sel"
                    )
                    max_view_week = 53 if year_has_week53(view_year) else 52
                    selected_view_weeks = st.multiselect(
                        "Weeks", list(range(1, max_view_week + 1)),
                        format_func=lambda w: f"Week {w}",
                        key="view_week_sel"
                    )
                    scraped_date_range = []
                    scraped_week_selection = (view_year, selected_view_weeks) if selected_view_weeks else None

                st.markdown("---")

                # ── Stay Dates (Checkin) ─────────────────────────────────────
                st.markdown("**Stay Dates**")
                stay_mode = st.radio("Stay date input mode", ["📅 Calendar", "📆 Week"],
                                    horizontal=True, key="stay_date_mode",
                                    label_visibility="collapsed")

                selected_stay_weeks = []  # always initialize

                if stay_mode == "📅 Calendar":
                    price_date_range = st.date_input(
                        "Stay Dates", value=[],
                        key="price_dates_unique",
                        label_visibility="collapsed"
                    )
                    stay_week_selection = None
                else:
                    stay_year_options = list(range(current_year - 2, current_year + 3))
                    stay_year = st.selectbox(
                        "Year", stay_year_options, index=2,
                        key="stay_year_sel"
                    )
                    max_stay_week = 53 if year_has_week53(stay_year) else 52
                    selected_stay_weeks = st.multiselect(
                        "Weeks", list(range(1, max_stay_week + 1)),
                        format_func=lambda w: f"Week {w}",
                        key="stay_week_sel"
                    )
                    price_date_range = []
                    stay_week_selection = (stay_year, selected_stay_weeks) if selected_stay_weeks else None
            
            with st.expander("🎨 Color Configuration", expanded=True):
                st.info("💡 Select a color preset for this location")
                
                available_presets = get_color_presets_for_location(location)
                
                if available_presets:
                    selected_preset = st.selectbox(
                        "Color Setup",
                        available_presets,
                        help="Color presets available for this location",
                        key="price_preset"
                    )
                    
                    if 'last_price_preset' not in st.session_state or \
                       st.session_state.last_price_preset != selected_preset:
                        
                        st.session_state.price_color_ranges = get_color_config_by_name(selected_preset)
                        st.session_state.last_price_preset = selected_preset
                    
                else:
                    st.warning(f"⚠️ No color presets for {location}")
                    st.session_state.price_color_ranges = get_default_color_ranges()['zone1']
            
            st.markdown("---")
            query_button = st.button("🚀 Execute Query", type="primary", use_container_width=True)

        if query_button:
            def weeks_to_date_range(year, weeks):
                import datetime as _dt
                all_days = []
                invalid = []
                for w in weeks:
                    try:
                        all_days += [
                            _dt.date.fromisocalendar(year, w, 1),
                            _dt.date.fromisocalendar(year, w, 7)
                        ]
                    except ValueError:
                        invalid.append(w)
                if invalid:
                    st.warning(f"⚠️ Week(s) {invalid} don't exist in {year} and were skipped.")
                return (min(all_days), max(all_days)) if all_days else (None, None)

            # ── Resolve scraped (view) dates ──────────────────────────────
            if view_mode == "📅 Calendar":
                if len(scraped_date_range) != 2:
                    st.error("⚠️ Please select both View Dates (start & end)")
                    st.stop()
                scraped_start = scraped_date_range[0].strftime("%Y-%m-%d")
                scraped_end   = scraped_date_range[1].strftime("%Y-%m-%d")
            else:
                if not selected_view_weeks:
                    st.error("⚠️ Please select at least one View week")
                    st.stop()
                _vy, _vw = scraped_week_selection
                _vs, _ve = weeks_to_date_range(_vy, _vw)
                if _vs is None:
                    st.error(f"❌ None of the selected View weeks exist in {_vy}.")
                    st.stop()
                scraped_start = _vs.strftime("%Y-%m-%d")
                scraped_end   = _ve.strftime("%Y-%m-%d")

            # ── Resolve stay (price/checkin) dates ───────────────────────
            if stay_mode == "📅 Calendar":
                if len(price_date_range) != 2:
                    st.error("⚠️ Please select both Stay Dates (start & end)")
                    st.stop()
                price_start = price_date_range[0].strftime("%d-%m-%Y")
                price_end   = price_date_range[1].strftime("%d-%m-%Y")
            else:
                if not selected_stay_weeks:
                    st.error("⚠️ Please select at least one Stay week")
                    st.stop()
                _sy, _sw = stay_week_selection
                _ss, _se = weeks_to_date_range(_sy, _sw)
                if _ss is None:
                    st.error(f"❌ None of the selected Stay weeks exist in {_sy}.")
                    st.stop()
                price_start = _ss.strftime("%d-%m-%Y")
                price_end   = _se.strftime("%d-%m-%Y")

            with st.spinner("🔍 Searching hotels..."):
                filters = {
                    "location": location,
                    "time": time_of_day,
                    "persons": str(persons),
                    "nights": str(nights)
                }
                date_range = f"{price_start} - {price_end}"
                results = query_hotels(
                    filters=filters,
                    date_range=date_range,
                    scraped_date_start=scraped_start,
                    scraped_date_end=scraped_end
                )

            if results:
                st.session_state.results = results
                st.success(f"✅ Found {len(results)} hotel records!")
            else:
                st.error("❌ No data found for your criteria")

        if 'results' in st.session_state and st.session_state.results:
            df = pd.DataFrame(st.session_state.results)
            df['price'] = pd.to_numeric(df['price'], errors='coerce')
            df = df.dropna(subset=['price'])

            if breakfast_filter and cancellation_filter:
                df = df[(df['breakfast_included'] == True) & (df['free_cancellation'] == True)]
                st.success(f"🍳✅ Filtered to {len(df)} records with breakfast included and free cancellation")
            elif breakfast_filter and not cancellation_filter:
                df = df[(df['breakfast_included'] == True) & (df['free_cancellation'] == False)]
                st.success(f"🍳 Filtered to {len(df)} records with breakfast included")
            elif cancellation_filter and not breakfast_filter:
                df = df[(df['breakfast_included'] == False) & (df['free_cancellation'] == True)]
                st.success(f"✅ Filtered to {len(df)} records with free cancellation")
            else:
                df = df[(df['breakfast_included'] == False) & (df['free_cancellation'] == False)]
            
            if not df.empty:
                st.markdown('<div class="hotel-selector">', unsafe_allow_html=True)

                col1, col2, col3 = st.columns([3, 1, 1])
                with col1:
                    st.markdown("### 🏨 Hotel Selection")
                    
                    unique_hotels = sorted(df['name'].unique())
                    
                    if 'selected_hotels' not in st.session_state:
                        st.session_state.selected_hotels = []
                    if 'multiselect_key' not in st.session_state:
                        st.session_state.multiselect_key = 0
                    
                    st.markdown("**Quick Selection:**")
                    location_zones = get_zones_for_location(location)
                    COLS_PER_ROW   = 6
                    total_btns     = 2 + len(location_zones)   # "Select All" + "Clear All" + N zones
                    btn_row        = st.columns(min(COLS_PER_ROW, total_btns))
                    btn_idx        = 0

                    # Select All (fixed, always first)
                    with btn_row[btn_idx % COLS_PER_ROW]:
                        if st.button("✅ Select All", key="select_all_btn", use_container_width=True):
                            st.session_state.selected_hotels = list(unique_hotels)
                            st.session_state.multiselect_key += 1
                    btn_idx += 1

                    # Clear All (fixed, always second)
                    with btn_row[btn_idx % COLS_PER_ROW]:
                        if st.button("❌ Clear All", key="clear_all_btn", use_container_width=True):
                            st.session_state.selected_hotels = []
                            st.session_state.multiselect_key += 1
                    btn_idx += 1

                    # One button per DB zone
                    for zone in location_zones:
                        z_name   = zone.get('zone_name', 'Zone')
                        z_hotels = zone.get('hotels', [])
                        # Unique key: zone_name + location prevents collisions across locations
                        btn_key  = f"zone_btn_{z_name}__{zone.get('location','')}"

                        if btn_idx > 0 and btn_idx % COLS_PER_ROW == 0:
                            btn_row = st.columns(COLS_PER_ROW)   # start a new row

                        with btn_row[btn_idx % COLS_PER_ROW]:
                            if st.button(z_name, key=btn_key, use_container_width=True):
                                st.session_state.selected_hotels = [h for h in z_hotels if h in unique_hotels]
                                st.session_state.multiselect_key += 1
                        btn_idx += 1

                    
                    default_hotels = [h for h in st.session_state.selected_hotels if h in unique_hotels]

                    hotels = st.multiselect(
                        "Select hotels to analyze:",
                        unique_hotels,
                        default=default_hotels,
                        key=f"hotel_selector_{st.session_state.multiselect_key}"
                    )
                    
                    st.session_state.selected_hotels = hotels

                with col2:
                    st.metric("Total Hotels", len(unique_hotels))
                    for zone in location_zones:
                        cnt = len([h for h in zone.get('hotels',[]) if h in unique_hotels])
                        if cnt > 0: st.metric(f"{zone['zone_name']} Available", cnt)

                with col3:
                    if hotels:
                        st.metric("Selected", len(hotels))
                        for zone in location_zones:
                            cnt = len([h for h in zone.get('hotels',[]) if h in hotels])
                            if cnt > 0: st.metric(f"{zone['zone_name']} Selected", cnt)

                st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('<div class="hotel-selector">', unsafe_allow_html=True)
                st.markdown("### 📈 Line Chart Hotel Selection (Optional)")

                if 'line_hotels' not in st.session_state:
                    st.session_state.line_hotels = []

                if 'show_rates_pricing' not in st.session_state:
                    st.session_state.show_rates_pricing = True

                all_hotels = sorted(df['name'].unique())
                valid_defaults = [h for h in st.session_state.get('line_hotels', []) if h in all_hotels]

                col_select, col_display = st.columns([3, 1])

                with col_select:
                    line_hotels = st.multiselect(
                        "Select hotels for line chart trend:",
                        options=all_hotels,
                        default=valid_defaults,
                        key="line_chart_selector"
                    )
                    st.session_state.line_hotels = line_hotels

                with col_display:
                    st.markdown("**Display Options**")
                    show_rates = st.checkbox(
                        "💰 Show rates and pricing",
                        value=st.session_state.get('show_rates_pricing', True),
                        key="show_rates_checkbox"
                    )
                    st.session_state.show_rates_pricing = show_rates

                    show_std = st.checkbox(
                        "📊 Show standardized comparison",
                        value=st.session_state.get('show_std_comparison', False),
                        key="show_std_checkbox"
                    )
                    st.session_state.show_std_comparison = show_std

                st.markdown('</div>', unsafe_allow_html=True)

                if hotels:
                    filtered_df = df[df['name'].isin(hotels)].copy()
                    filtered_df['price_date'] = pd.to_datetime(filtered_df['price_date'], format='%Y-%m-%d')

                    # Bar chart data
                    bar_avg = filtered_df.groupby('price_date')['price'].mean().reset_index()
                    bar_avg['week_num'] = bar_avg['price_date'].dt.isocalendar().week
                    bar_avg['x_label'] = make_x_label(bar_avg['price_date'])

                    price_color_ranges = st.session_state.get('price_color_ranges', get_default_color_ranges()['zone1'])
                    bar_avg['bar_color'] = bar_avg['price'].apply(
                        lambda x: get_color_from_price_ranges(x, price_color_ranges)
                    )

                    def add_week_annotations(fig, bar_avg):
                        """Add week label bands below the x-axis spanning the full week."""
                        x_labels = bar_avg['x_label'].tolist()

                        for week_num, week_group in bar_avg.groupby('week_num'):
                            x_indices = [x_labels.index(lbl) for lbl in week_group['x_label'] if lbl in x_labels]
                            if not x_indices:
                                continue

                            x_start = min(x_indices) - 0.45
                            x_end = max(x_indices) + 0.45
                            mid_x = (x_start + x_end) / 2
                            bg_color = st.get_option("theme.backgroundColor") or "#0e1117"

                            fig.add_shape(
                                type="rect",
                                xref="x",
                                yref="paper",
                                x0=x_start,
                                x1=x_end,
                                y0=-0.27,
                                y1=-0.18,
                                fillcolor=bg_color,
                                line=dict(color="white", width=2),
                                layer="above"
                            )

                            fig.add_annotation(
                                x=mid_x,
                                y=-0.25,
                                xref="x",
                                yref="paper",
                                text=f"<b>Week {week_num}</b>",
                                showarrow=False,
                                font=dict(size=12, color="white"),
                                align="center",
                            )

                        return fig
                    
                    if 'std_top_value' not in st.session_state:
                        st.session_state.std_top_value = get_std_top_value()
                    
                    show_rates_val = st.session_state.show_rates_pricing
                    show_std_val = st.session_state.show_std_comparison
                    std_top = st.session_state.std_top_value    

                    # Build line_avg if line hotels selected
                    line_avg = None
                    if line_hotels:
                        line_df = df[df['name'].isin(line_hotels)].copy()
                        line_df['price'] = pd.to_numeric(line_df['price'], errors='coerce')
                        pivot_line = line_df.pivot_table(
                            index='scrape_date', columns='price_date', values='price', aggfunc='mean'
                        )
                        line_avg = pivot_line.mean(axis=0).reset_index()
                        line_avg.columns = ['price_date', 'price']
                        line_avg = line_avg.dropna(subset=['price'])
                        line_avg['price_date'] = pd.to_datetime(line_avg['price_date'])
                        line_avg['x_label'] = make_x_label(line_avg['price_date'])

                    def build_bar_fig(title, show_labels=True, yaxis_range=None):
                        fig = px.bar(
                            bar_avg, x='x_label', y='price',
                            labels={'price': 'Average Price (€)', 'x_label': 'Date'},
                            title=title
                        )
                        fig.data[0].marker.color = bar_avg['bar_color'].tolist()

                        if show_labels:
                            fig.data[0].text = bar_avg['price'].apply(lambda x: f'€{x:.1f}')
                            fig.data[0].textposition = 'outside'
                            fig.data[0].hovertemplate = '<b>%{x}</b><br>Price: €%{y:.2f}<extra></extra>'
                        else:
                            fig.data[0].text = None
                            fig.data[0].textposition = 'none'
                            fig.data[0].hovertemplate = '<b>%{x}</b><extra></extra>'

                        # Scatter dots only if line hotels selected AND labels shown
                        if line_avg is not None and show_labels:
                            for idx, line_row in line_avg.iterrows():
                                matching_bar = bar_avg[bar_avg['price_date'] == line_row['price_date']]
                                if not matching_bar.empty:
                                    fig.add_scatter(
                                        x=[matching_bar.iloc[0]['x_label']], y=[line_row['price']],
                                        mode='markers', name='Trend',
                                        marker=dict(color='red', size=14),
                                        showlegend=(idx == 0),
                                        hovertemplate='<b>Trend</b><br>Price: €%{y:.2f}<extra></extra>'
                                    )

                        layout = dict(
                            height=550,
                            xaxis_title="",
                            yaxis_title="Average Price (€)" if show_labels else "",
                            xaxis=dict(tickangle=0, tickfont=dict(size=11)),
                            hovermode='x unified',
                            margin=dict(b=110),
                            showlegend=show_labels and line_avg is not None
                        )
                        if yaxis_range is not None:
                            layout['yaxis'] = dict(range=[0, yaxis_range])

                        fig.update_layout(**layout)

                        if not show_labels:
                            fig.update_yaxes(showticklabels=False, showgrid=False)
                            fig.update_xaxes(showgrid=False)

                        fig = add_week_annotations(fig, bar_avg)
                        return fig

                    # All 4 cases handled
                    if show_rates_val and show_std_val:
                        st.plotly_chart(build_bar_fig(f'Average Prices Across Selected Hotels', show_labels=True, yaxis_range=std_top), use_container_width=True)
                    elif show_rates_val and not show_std_val:
                        st.plotly_chart(build_bar_fig('Average Prices Across Selected Hotels', show_labels=True), use_container_width=True)
                    elif show_std_val and not show_rates_val:
                        st.plotly_chart(build_bar_fig(f'Average Prices Across Selected Hotels', show_labels=False, yaxis_range=std_top), use_container_width=True)
                    else:
                        # Both unchecked — bars with no labels (existing behaviour)
                        st.plotly_chart(build_bar_fig('Average Prices Across Selected Hotels', show_labels=False), use_container_width=True)
                    
                    # Detailed table section
                    st.markdown("### 📋 Detailed Price Matrix")

                    pivot = filtered_df.pivot_table(
                        index=['scrape_date', 'name'], 
                        columns='price_date',
                        values='price', 
                        aggfunc='mean'
                    )

                    pivot.columns = [col.strftime('%Y-%m-%d') if isinstance(col, pd.Timestamp) else col for col in pivot.columns]
                    pivot = pivot.reset_index()

                    numeric_cols = [col for col in pivot.columns if col not in ['scrape_date', 'name', 'breakfast_included']]

                    unique_dates = pivot['scrape_date'].unique()
                    if len(unique_dates) > 1:
                        final_pivot = pd.DataFrame()
                        for i, date in enumerate(unique_dates):
                            group = pivot[pivot['scrape_date'] == date]
                            final_pivot = pd.concat([final_pivot, group], ignore_index=True)
                            if i < len(unique_dates) - 1:
                                empty_row = {col: None for col in pivot.columns}
                                empty_df = pd.DataFrame([empty_row])
                                final_pivot = pd.concat([final_pivot, empty_df], ignore_index=True)
                        pivot = final_pivot

                    if numeric_cols:
                        valid_data = pivot[pivot['scrape_date'].notnull()]
                        averages = valid_data[numeric_cols].mean()

                        empty_row = {col: None for col in pivot.columns}
                        empty_df = pd.DataFrame([empty_row])
                        pivot = pd.concat([pivot, empty_df], ignore_index=True)

                        avg_row = {'scrape_date': 'AVERAGE', 'name': 'AVERAGE'}
                        for col in numeric_cols:
                            if not pd.isna(averages[col]):
                                avg_row[col] = round(averages[col], 2)
                            else:
                                avg_row[col] = None
                        avg_df = pd.DataFrame([avg_row])
                        pivot = pd.concat([pivot, avg_df], ignore_index=True)

                    gb = GridOptionsBuilder.from_dataframe(pivot)
                    gb.configure_columns(['scrape_date', 'name'], pinned='left', minWidth=150)
                    gb.configure_columns(
                        numeric_cols,
                        type=['numericColumn'],
                        precision=2,
                        minWidth=100,
                        maxWidth=100
                    )
                    gb.configure_default_column(resizable=True)
                    gb.configure_grid_options(rowHeight=35)
                    gridOptions = gb.build()

                    AgGrid(
                        pivot,
                        gridOptions=gridOptions,
                        height=600,
                        fit_columns_on_grid_load=False,
                        enable_enterprise_modules=False,
                    )
                
                else:
                    st.info("👆 Please select one or more hotels to view the analysis")
            
            else:
                st.error("❌ No valid price data found")

        else:
            st.markdown("""
            <div style="text-align: center; padding: 3rem; background: #f8f9fa; border-radius: 15px; margin: 2rem 0;">
                <h2>🎯 Welcome to Hotel Price Analytics</h2>
                <p style="font-size: 1.2em; color: #666; margin-bottom: 2rem;">
                    Configure your search parameters in the sidebar and click "Execute Query" to begin analysis
                </p>
            </div>
            """, unsafe_allow_html=True)

if tab2:
# ==================== TAB 2: CALENDAR HEATMAP ====================
    with tab2:
        st.markdown("""
        <div class="main-header">
            <h1>📅 Weekly Calendar Heatmap</h1>
            <p>Hotel availability and pricing trends by week</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.sidebar:
            st.markdown("### 📅 Calendar Configuration")

            with st.expander("📍 Location & Zone Selection", expanded=True):
                allowed_locations  = st.session_state.get("locations", [])
                cal_allowed        = [l for l in allowed_locations if l == "tampere"] or allowed_locations
                calendar_location  = st.selectbox("Select Location", cal_allowed,
                                                   index=0 if cal_allowed else None,
                                                   key="calendar_location_key")
                # Build zone dropdown from DB for the selected location
                cal_zones = get_zones_for_location(calendar_location) if calendar_location else []
                if cal_zones:
                    zone_selection = st.selectbox(
                        "Select Zone",
                        [z['zone_name'] for z in cal_zones],
                        key="zone_selection_key"
                    )
                else:
                    st.warning(f"⚠️ No zones found for {calendar_location}")
            
            with st.expander("📅 Date Range", expanded=True):
                selected_cal_weeks = [] 

                cal_mode = st.radio("Date input mode", ["📅 Calendar", "📆 Week"],
                                    horizontal=True, key="cal_date_mode",
                                    label_visibility="collapsed")

                if cal_mode == "📅 Calendar":
                    calendar_start = st.date_input("Start Date", value=datetime(2025, 12, 1), key="cal_start_key")
                    calendar_end = st.date_input("End Date", value=datetime(2025, 12, 15), key="cal_end_key")
                else:
                    current_year = datetime.now().year
                    cal_year_options = list(range(current_year - 2, current_year + 3))
                    cal_year = st.selectbox(
                        "Year", cal_year_options,
                        index=cal_year_options.index(current_year),
                        key="cal_year_sel"
                    )

                    # Only show week 53 if the selected year actually has it
                    import datetime as _dt
                    def year_has_week53(year):
                        try:
                            _dt.date.fromisocalendar(year, 53, 1)
                            return True
                        except ValueError:
                            return False

                    max_week = 53 if year_has_week53(cal_year) else 52
                    all_cal_weeks = list(range(1, max_week + 1))

                    selected_cal_weeks = st.multiselect(
                        "Weeks", all_cal_weeks,
                        format_func=lambda w: f"Week {w}",
                        key="cal_week_sel"
                    )

                    if selected_cal_weeks:
                        all_days = []
                        invalid_weeks = []
                        for w in selected_cal_weeks:
                            try:
                                monday = _dt.date.fromisocalendar(cal_year, w, 1)
                                sunday = _dt.date.fromisocalendar(cal_year, w, 7)
                                all_days += [monday, sunday]
                            except ValueError:
                                invalid_weeks.append(w)

                        if invalid_weeks:
                            st.warning(f"⚠️ Week(s) {invalid_weeks} don't exist in {cal_year} and were skipped.")

                        if all_days:
                            calendar_start = min(all_days)
                            calendar_end = max(all_days)
                        else:
                            calendar_start = None
                            calendar_end = None
                            st.error(f"None of the selected weeks exist in {cal_year}.")
                    else:
                        calendar_start = None
                        calendar_end = None

            
            
            with st.expander("🎨 Color Configuration", expanded=True):
                st.info("💡 Select a color preset for this location")
                
                available_presets = get_color_presets_for_location(calendar_location)
                
                if available_presets:
                    selected_calendar_preset = st.selectbox(
                        "Color Setup",
                        available_presets,
                        help="Color presets available for this location",
                        key="calendar_preset"
                    )
                    
                    if 'last_calendar_preset' not in st.session_state or \
                       st.session_state.last_calendar_preset != selected_calendar_preset:
                        
                        st.session_state.color_ranges = get_color_config_by_name(selected_calendar_preset)
                        st.session_state.last_calendar_preset = selected_calendar_preset
                    
                else:
                    st.warning(f"⚠️ No color presets for {calendar_location}")
                    st.session_state.color_ranges = get_default_color_ranges()['zone1']
                
            st.markdown("---")
            calendar_query_button = st.button("🔄 Load Calendar Data", type="primary", use_container_width=True)

        if calendar_query_button:
            if cal_mode == "📆 Week" and (not selected_cal_weeks or calendar_start is None or calendar_end is None):
                st.error("⚠️ Please select at least one valid week before loading data.")
                st.stop()
            elif cal_mode == "📅 Calendar" and (calendar_start is None or calendar_end is None):
                st.error("⚠️ Please select a valid date range.")
                st.stop()

            with st.spinner("📊 Generating calendar data..."):
                st.session_state.calendar_data = query_calendar_data(
                    price_start_date=calendar_start,
                    price_end_date=calendar_end,
                    zone_filter=zone_selection,
                    location=calendar_location
                )
                st.session_state.calendar_date_range = (calendar_start, calendar_end)
                st.session_state.calendar_zone = zone_selection
                st.session_state.calendar_location = calendar_location
        
        if 'calendar_data' in st.session_state:
            metrics = st.session_state.calendar_data
            
            df_cal_availability = pd.DataFrame({
                'date': list(metrics['availability'].keys()),
                'value': list(metrics['availability'].values())
            })
            
            df_cal_price = pd.DataFrame({
                'date': list(metrics['price_avg'].keys()),
                'value': list(metrics['price_avg'].values())
            })
            
            df_cal_free_cancel = pd.DataFrame({
                'date': list(metrics['free_cancel_avg'].keys()),
                'value': list(metrics['free_cancel_avg'].values())
            })
            
            for df in [df_cal_availability, df_cal_price, df_cal_free_cancel]:
                df['date'] = pd.to_datetime(df['date'])
                df['year'] = df['date'].dt.year
                df['week'] = df['date'].dt.isocalendar().week
                df['day_name'] = df['date'].dt.strftime('%A')
                df['date_str'] = df['date'].dt.strftime('%m/%d/%Y')
                df['day_num'] = df['date'].dt.dayofweek
            
            all_years = sorted(set(
                list(df_cal_availability['year'].unique()) + 
                list(df_cal_price['year'].unique()) + 
                list(df_cal_free_cancel['year'].unique())
            ))
            all_weeks = sorted(set(
                list(df_cal_availability['week'].unique()) + 
                list(df_cal_price['week'].unique()) + 
                list(df_cal_free_cancel['week'].unique())
            ))
            
            with st.expander("📊 Display Metric", expanded=True):
                color_metric = st.selectbox(
                    "Display Metric By",
                    ["availability", "price_avg", "free_cancel_avg"],
                    format_func=lambda x: {"availability": "Hotel Availability %", "price_avg": "Average Price (€)", "free_cancel_avg": "Free Cancellation Avg (€)"}.get(x),
                    key="color_metric_key"
                )
            
            col1, col2 = st.columns(2)
            with col1:
                selected_years = st.multiselect(
                    "Select Years",
                    all_years,
                    default=all_years,
                    key="selected_years_key"
                )
            
            with col2:
                selected_weeks = st.multiselect(
                    "Select Weeks",
                    all_weeks,
                    default=all_weeks,
                    key="selected_weeks_key"
                )
            if not selected_years or not selected_weeks:
                st.error("Please select at least one year and one week")
            else:
                metric_map = {
                    'availability': df_cal_availability,
                    'price_avg': df_cal_price,
                    'free_cancel_avg': df_cal_free_cancel
                }
                df_cal_display = metric_map[color_metric].copy()
                
                filtered_cal_display = df_cal_display[
                    (df_cal_display['year'].isin(selected_years)) & 
                    (df_cal_display['week'].isin(selected_weeks))
                ].copy()
                
                if filtered_cal_display.empty:
                    st.warning("No data available for selected filters")
                else:
                    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                    
                    html = '<table class="calendar-table"><tr>'
                    html += '<td class="week-label"></td>'
                    for day in day_order:
                        html += f'<td class="day-label">{day}</td>'
                    html += '</tr>'
                    
                    week_year_pairs = []
                    for week in sorted(selected_weeks):
                        for year in sorted(selected_years):
                            week_year_pairs.append((week, year))
                    
                    for week, year in week_year_pairs:
                        week_data_display = filtered_cal_display[
                            (filtered_cal_display['week'] == week) & 
                            (filtered_cal_display['year'] == year)
                        ].copy()
                        
                        html += '<tr>'
                        html += f'<td class="week-label">Week {week} - {year}</td>'
                        
                        for day in day_order:
                            day_data_display = week_data_display[week_data_display['day_name'] == day]
                            
                            if not day_data_display.empty:
                                row_display = day_data_display.iloc[0]
                                display_value = row_display['value']
                                date_str = datetime.strptime(row_display['date_str'], "%m/%d/%Y").strftime("%d/%m/%Y")
                                
                                bg_color = get_color_from_price_ranges(display_value, st.session_state.color_ranges)
                                text_color = get_text_color_from_background(bg_color)
                                
                                if color_metric == "availability":
                                    display_text = f"{display_value:.1f}%"
                                else:
                                    display_text = f"€{display_value:.2f}"
                                
                                html += f'''<td style="background-color: {bg_color};">
                                    <div class="cell-content" style="color: {text_color};">
                                        <div class="cell-date">{date_str}</div>
                                        <div class="cell-value">{display_text}</div>
                                    </div>
                                </td>'''
                            else:
                                html += '<td class="empty-cell"></td>'
                        
                        html += '</tr>'
                    
                    html += '</table>'
                    st.markdown(html, unsafe_allow_html=True)
                    
                    st.markdown("---")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.info(f"**Zone:** {zone_selection.replace('_', ' ').title()}")
                    with col2:
                        st.info(f"**Display:** {color_metric.replace('_', ' ').title()}")
                    with col3:
                        st.info(f"**Location:** {st.session_state.calendar_location.title()}")
                    with col4:
                        st.info(f"**Color Range:** {filtered_cal_display['value'].min():.1f} - {filtered_cal_display['value'].max():.1f}")
        
        else:
            st.info("👈 Configure settings and click 'Load Calendar Data' to begin")

if tab_matrix:
    with tab_matrix:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("""
            <div class="main-header">
                <h1>📤 Matrix Automation</h1>
                <p>Generate and email hotel price matrices on demand</p>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="user-info">
                <p><strong>👤 Logged in as:</strong><br>
                {st.session_state.get('authenticated_user', 'Unknown')}</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("🚪 Logout", key="matrix_logout",
                         type="secondary", use_container_width=True):
                logout()

        st.markdown("### ⚙️ Configuration")
        fc1, fc2, fc3 = st.columns(3)

        with fc1:
            st.markdown("**📍 Location & Zone**")
            mx_location = st.selectbox("Location", ALL_LOCATIONS, key="mx_location")
            mx_zones_raw = get_zones_for_location(mx_location)
            if mx_zones_raw:
                mx_zone = st.selectbox(
                    "Zone",
                    [z["zone_name"] for z in mx_zones_raw],
                    key="mx_zone"
                )
            else:
                st.warning(
                    f"No zones for **{mx_location}**. "
                    "Create one in Admin → Zone Management."
                )
                mx_zone = ""

        with fc2:
            st.markdown("**📅 Date & Query Options**")
            mx_start = st.date_input(
                "Start Date", value=datetime.today(), key="mx_start",
                help="First check-in date to include. Scrape date = this date."
            )
            mx_days = st.number_input(
                "Days Forward",
                min_value=1, max_value=365, value=365,
                step=1, key="mx_days"
            )
            mx_time = st.selectbox(
                "Scrape Time",
                ["morning", "evening"],
                key="mx_time",
                help="Morning = AM scrape, Evening = PM scrape."
            )
            mx_persons = st.selectbox(
                "Persons",
                [1, 2],
                index=1,
                key="mx_persons",
                format_func=lambda x: f"{x} Person{'s' if x > 1 else ''}"
            )
            st.markdown("**🔽 Data Filters**")
            mx_filter_mode = st.radio(
                "Include in Excel",
                ["All data (no filter)", "Apply filters"],
                key="mx_filter_mode",
                horizontal=True
            )
            if mx_filter_mode == "Apply filters":
                mx_breakfast = st.checkbox("🍳 Breakfast Included", value=False, key="mx_breakfast")
                mx_free_cancel = st.checkbox("✅ Free Cancellation", value=False, key="mx_free_cancel")
            else:
                mx_breakfast = None
                mx_free_cancel = None

        with fc3:
            st.markdown("**📧 Email Recipients**")
            saved_emails = get_saved_emails()
            if saved_emails:
                mx_selected_emails = st.multiselect(
                    "Select from saved emails",
                    options=saved_emails,
                    key="mx_email_select"
                )
            else:
                st.info("No saved emails yet — add them in Admin → Email Management.")
                mx_selected_emails = []

            mx_extra_emails_raw = st.text_area(
                "Extra recipients (one per line)",
                placeholder="extra@example.com",
                height=80,
                key="mx_extra_emails"
            )

        st.markdown("---")
        mx_run = st.button(
            "🚀 Generate & Send Matrix",
            type="primary",
            use_container_width=True,
            key="mx_run_btn"
        )

        if mx_run:
            extra_emails   = [e.strip() for e in mx_extra_emails_raw.splitlines()
                               if e.strip()]
            all_recipients = list(dict.fromkeys(mx_selected_emails + extra_emails))

            errs = []
            if not mx_zone:
                errs.append("Select a zone.")
            if not all_recipients:
                errs.append("Add at least one email recipient.")
            for err in errs:
                st.error(err)

            if not errs:
                start_dt = datetime(mx_start.year, mx_start.month, mx_start.day)

                # ── Step 1: Query DynamoDB ────────────────────────────────
                with st.status("🔍 Querying hotel prices…", expanded=True) as status:
                    all_rows = []
                    st.write(
                        f"  › {mx_persons} person{'s' if mx_persons > 1 else ''} "
                        f"/ {mx_time} scrape …"
                    )
                    rows = _query_matrix_data(
                        location=mx_location,
                        persons=mx_persons,
                        time_val=mx_time,
                        start_date=start_dt,
                        days_forward=int(mx_days)
                    )
                    all_rows.extend(rows)
                    st.write(f"    ✓ {len(rows):,} records")
                    status.update(
                        label=f"✅ {len(all_rows):,} records fetched",
                        state="complete"
                    )

                if not all_rows:
                    st.error(
                        "❌ No data found for the selected scrape date and options. "
                        "Check the scraper has run on this date for this location and time."
                    )
                else:
                    df_raw = pd.DataFrame(all_rows)
                    df_raw["price"] = pd.to_numeric(df_raw["price"], errors="coerce")
                    df_raw = df_raw.dropna(subset=["price"])

                    # ── Step 2: Filter by zone ────────────────────────────
                    zone_hotels       = _resolve_zone_hotels(mx_zone, mx_location)
                    df_zone           = df_raw[df_raw["name"].isin(zone_hotels)].copy()
                    total_zone_hotels = len(zone_hotels)
                    found_hotels      = df_zone["name"].nunique()

                    st.info(
                        f"🏨 **{found_hotels} / {total_zone_hotels}** zone hotels "
                        f"have data.  "
                        f"Total price records: **{len(df_zone):,}**."
                    )

                    if df_zone.empty:
                        st.error(
                            "❌ None of the zone hotels have data "
                            "for the selected scrape date."
                        )
                    else:
                        # ── Step 3: Apply breakfast/cancellation filter ───
                        if mx_filter_mode == "Apply filters":
                            df_excel = df_zone.copy()

                            if mx_breakfast and mx_free_cancel:
                                df_excel = df_excel[
                                    (df_excel["breakfast_included"] == True) &
                                    (df_excel["free_cancellation"] == True)
                                ]
                                filter_desc = "Breakfast_FreeCancel"
                            elif mx_breakfast and not mx_free_cancel:
                                df_excel = df_excel[
                                    (df_excel["breakfast_included"] == True) &
                                    (df_excel["free_cancellation"] == False)
                                ]
                                filter_desc = "Breakfast"
                            elif mx_free_cancel and not mx_breakfast:
                                df_excel = df_excel[
                                    (df_excel["breakfast_included"] == False) &
                                    (df_excel["free_cancellation"] == True)
                                ]
                                filter_desc = "FreeCancel"
                            else:
                                # Both unchecked — no extras
                                df_excel = df_excel[
                                    (df_excel["breakfast_included"] == False) &
                                    (df_excel["free_cancellation"] == False)
                                ]
                                filter_desc = "NoExtras"

                            if df_excel.empty:
                                st.error("❌ No records match the selected filters. Try different filter options.")
                                st.stop()

                            st.info(f"🔽 Filter applied — **{filter_desc.replace('_', ' + ')}**: {len(df_excel):,} records")

                        else:
                            df_excel = df_zone.copy()
                            filter_desc = "all"

                        # ── Step 4: Build Excel ───────────────────────────
                        with st.spinner("📊 Building Excel matrix…"):
                            xlsx_bytes = _build_excel_matrix(
                                df=df_excel,
                                zone_name=mx_zone,
                                location=mx_location,
                                persons_list=[mx_persons],
                                single_sheet=filter_desc.replace("_", " + ") if mx_filter_mode == "Apply filters" else None,
                            )

                        persons_str = f"{mx_persons}p"
                        filename    = (
                            f"price_matrix_{mx_location}_"
                            f"{mx_zone.replace(' ', '_')}_"
                            f"{mx_start.strftime('%Y%m%d')}_"
                            f"{mx_time}_"
                            f"{filter_desc}.xlsx"
                        )
                        subject = (
                            f"Hotel Price Matrix – {mx_location.title()} | "
                            f"{mx_zone} | {persons_str} | "
                            f"{mx_time.title()} | "
                            f"{mx_start.strftime('%d/%m/%Y')}"
                        )
                        body = (
                            f"Please find attached the hotel price matrix.\n\n"
                            f"Location  : {mx_location.title()}\n"
                            f"Zone      : {mx_zone}\n"
                            f"Persons   : {persons_str}\n"
                            f"Scrape    : {mx_time}\n"
                            f"Start date: {mx_start.strftime('%d/%m/%Y')}\n"
                            f"Days fwd  : {int(mx_days)}\n"
                            f"Filter    : {filter_desc.replace('_', ' + ')}\n"
                            f"Hotels    : {found_hotels} of "
                            f"{total_zone_hotels} in zone\n"
                            f"Generated : "
                            f"{datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
                            f"Sent from the Hotel Dashboard."
                        )

                        # ── Step 5: Send email ────────────────────────────
                        with st.spinner(
                            f"📧 Sending to {len(all_recipients)} recipient(s)…"
                        ):
                            try:
                                _send_matrix_email(
                                    recipients=all_recipients,
                                    subject=subject,
                                    body=body,
                                    xlsx_bytes=xlsx_bytes,
                                    filename=filename
                                )
                                st.success(
                                    f"✅ Matrix emailed to: "
                                    f"**{', '.join(all_recipients)}**"
                                )
                            except Exception as email_err:
                                st.error(f"❌ Email send failed: {email_err}")
                                st.warning("Download the file below.")

                        # Always offer local download as fallback
                        st.download_button(
                            label="⬇️ Download Matrix Excel",
                            data=xlsx_bytes,
                            file_name=filename,
                            mime=(
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            ),
                            use_container_width=True
                        )

if admin_panel:
    with admin_panel:

        try:
            users = table_user.scan().get("Items", [])
        except Exception as e:
            st.error(f"Failed to load users: {e}")
            users = []

        # ---- 1. Download Login Logs ----
        with st.expander("📥 Download Login Logs", expanded=False):
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                log_start = st.date_input("Start Date", key="log_start")
            with col2:
                log_end = st.date_input("End Date", key="log_end")
            with col3:
                st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                download_logs = st.button("Download Excel", use_container_width=True)

            if download_logs:
                if log_start > log_end:
                    st.error("Start date must be before end date")
                else:
                    with st.spinner("Preparing login logs..."):
                        try:
                            response = table_logs.scan(
                                FilterExpression=Attr("login_date").between(
                                    log_start.strftime("%Y-%m-%d"),
                                    log_end.strftime("%Y-%m-%d")
                                )
                            )
                            items = response.get("Items", [])
                            while "LastEvaluatedKey" in response:
                                response = table_logs.scan(
                                    FilterExpression=Attr("login_date").between(
                                        log_start.strftime("%Y-%m-%d"),
                                        log_end.strftime("%Y-%m-%d")
                                    ),
                                    ExclusiveStartKey=response["LastEvaluatedKey"]
                                )
                                items.extend(response.get("Items", []))

                            if not items:
                                st.warning("No login logs found for selected range")
                            else:
                                df_logs = pd.DataFrame(items).sort_values("login_ts", ascending=False)
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
                                    df_logs.to_excel(writer, index=False, sheet_name="Login Logs")
                                st.download_button(
                                    label="📥 Download Login Logs",
                                    data=buffer.getvalue(),
                                    file_name=f"login_logs_{log_start}_{log_end}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        except Exception as e:
                            st.error(f"Failed to generate logs: {e}")

        # ---- 2. Member Accounts ----
        with st.expander("👥 Member Accounts", expanded=False):

            st.markdown("#### ➕ Add Member Account")
            with st.form("add_staff_form"):
                new_username = st.text_input("User name")
                new_password = st.text_input("Password", type="password")
                new_locations = st.multiselect(
                    "Locations",
                    ["tampere", "oulu", "rauma", "turku", "jyvaskyla", "vaasa", "seinajoki"]
                )
                new_boards = st.multiselect(
                    "Boards Access",
                    AVAILABLE_BOARDS,
                    format_func=lambda x: {
                        "price_dashboard": "📊 Price Dashboard",
                        "historical_calendar": "📅 Historical Price Calendar",
                        "matrix_automation": "📤 Matrix Automation"
                    }.get(x, x)
                )
                create_btn = st.form_submit_button("Create account")

            if create_btn:
                if not new_username or not new_password:
                    st.error("Username and password are required")
                else:
                    try:
                        existing = table_user.get_item(Key={"username": new_username})
                        if "Item" in existing:
                            st.error("User already exists")
                        else:
                            table_user.put_item(Item={
                                "username": new_username,
                                "password": new_password,
                                "access": "basic",
                                "last_login": "",
                                "locations": new_locations,
                                "boards": new_boards
                            })
                            st.success("Member account created")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Failed to create user: {e}")

            # ── Active members ────────────────────────────────────────────
            st.markdown("#### 🟢 Active members")
            for user in users:
                if user.get("access") == "admin":
                    continue

                col1, col2, col3, col4, col5, col6, col7 = st.columns([2.5, 2.5, 2.5, 2.5, 2.5, 1.5, 1.5])

                with col1:
                    st.text_input("User name", value=user["username"], disabled=True, key=f"username_{user['username']}")
                with col2:
                    new_password = st.text_input("Password", value=user.get("password", ""), key=f"password_{user['username']}")
                with col3:
                    st.text_input("Last log in", value=user.get("last_login", ""), disabled=True, key=f"last_login_{user['username']}")
                with col4:
                    edit_locations = st.multiselect(
                        "Locations",
                        ["tampere", "oulu", "rauma", "turku", "jyvaskyla", "vaasa", "seinajoki"],
                        default=user.get("locations", []),
                        key=f"locations_{user['username']}"
                    )
                with col5:
                    edit_boards = st.multiselect(
                        "Boards",
                        AVAILABLE_BOARDS,
                        default=user.get("boards", []),
                        format_func=lambda x: {
                            "price_dashboard": "📊 Price Dashboard",
                            "historical_calendar": "📅 Historical Price Calendar",
                            "matrix_automation": "📤 Matrix Automation"
                        }.get(x, x),
                        key=f"boards_{user['username']}"
                    )
                with col6:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                    if st.button("Save", key=f"save_{user['username']}"):
                        try:
                            table_user.update_item(
                                Key={"username": user["username"]},
                                UpdateExpression="SET password = :p, locations = :l, boards = :b",
                                ExpressionAttributeValues={
                                    ":p": new_password,
                                    ":l": edit_locations,
                                    ":b": edit_boards
                                }
                            )
                            st.success(f"Updated {user['username']}")
                        except Exception as e:
                            st.error(f"Failed to update user: {e}")
                with col7:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                    if st.button("Delete", key=f"delete_{user['username']}", use_container_width=True):
                        st.session_state[f"delete_confirm_{user['username']}"] = True

                if st.session_state.get(f"delete_confirm_{user['username']}", False):
                    st.warning(f"⚠️ Are you sure you want to delete **{user['username']}**?")
                    col_c1, col_c2, _ = st.columns([1, 1, 2])
                    with col_c1:
                        if st.button("✅ Yes, Delete", key=f"confirm_delete_{user['username']}", use_container_width=True):
                            try:
                                table_user.delete_item(Key={"username": user["username"]})
                                st.success(f"User **{user['username']}** deleted")
                                st.session_state[f"delete_confirm_{user['username']}"] = False
                                time.sleep(1)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to delete user: {e}")
                    with col_c2:
                        if st.button("❌ Cancel", key=f"cancel_delete_{user['username']}", use_container_width=True):
                            st.session_state[f"delete_confirm_{user['username']}"] = False
                            st.rerun()
                    st.divider()

            # ── Admin accounts ────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 🔐 Admin Accounts")
            st.caption("Change admin password here. Admin accounts cannot be deleted.")

            for user in users:
                if user.get("access") != "admin":
                    continue

                ac1, ac2, ac3 = st.columns([3, 3, 1])
                with ac1:
                    st.text_input(
                        "Username",
                        value=user["username"],
                        disabled=True,
                        key=f"admin_username_{user['username']}"
                    )
                with ac2:
                    admin_new_pw = st.text_input(
                        "Password",
                        value=user.get("password", ""),
                        type="password",
                        key=f"admin_password_{user['username']}"
                    )
                with ac3:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                    if st.button("Save", key=f"admin_save_{user['username']}", use_container_width=True):
                        try:
                            table_user.update_item(
                                Key={"username": user["username"]},
                                UpdateExpression="SET password = :p",
                                ExpressionAttributeValues={":p": admin_new_pw}
                            )
                            st.success(f"✅ Password updated for **{user['username']}**")
                        except Exception as e:
                            st.error(f"Failed to update: {e}")
                st.divider()

        # ---- 3. Color Configuration ----
        with st.expander("🎨 Color Configuration Management", expanded=False):

            try:
                response = table_color.scan()
                color_configs = response.get('Items', [])
                while 'LastEvaluatedKey' in response:
                    response = table_color.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
                    color_configs.extend(response.get('Items', []))
            except Exception as e:
                st.error(f"Failed to load color configs: {e}")
                color_configs = []

            # ---- EXISTING CONFIGS ----
            if st.checkbox("Show existing configurations", value=False, key="show_existing_configs"):
                if color_configs:
                    for config in color_configs:
                        cid = config["id"]
                        state_key = f"edit_ranges_{cid}"

                        col1, col2, col3, col4, col5, col6, col7 = st.columns([1.5, 2, 2, 2, 1.5, 1, 1])
                        with col1: st.text(f"📝 {config.get('color_config_name', 'N/A')}")
                        with col2:
                            locations = ', '.join(config.get('locations', [])[:2])
                            if len(config.get('locations', [])) > 2:
                                locations += f" +{len(config['locations'])-2}"
                            st.text(f"📍 {locations}")
                        with col3: st.text(f"📊 {', '.join(config.get('dashboards', []))}")
                        with col4: st.text(f"🎨 {len(config.get('ranges', []))}")

                        with col5:
                            if st.button("✏️ Edit", key=f"edit_config_{cid}", use_container_width=True):
                                st.session_state[f"editing_config_{cid}"] = True

                        with col6:
                            if st.button("📋", key=f"copy_config_{cid}", use_container_width=True):
                                try:
                                    new_name = f"{config['color_config_name']}_copy"
                                    table_color.put_item(Item={
                                        'id': str(uuid.uuid4()),
                                        'color_config_name': new_name,
                                        'locations': config.get('locations', []),
                                        'dashboards': config.get('dashboards', []),
                                        'ranges': config.get('ranges', []),
                                        'created_at': datetime.now().isoformat(),
                                        'created_by': st.session_state.get('authenticated_user', 'admin')
                                    })
                                    st.success(f"Copied as '{new_name}'")
                                    time.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Failed to copy: {e}")

                        with col7:
                            if st.button("🗑️", key=f"delete_config_{cid}", use_container_width=True):
                                try:
                                    table_color.delete_item(Key={'color_config_name': config['color_config_name']})
                                    st.success("Deleted!")
                                    time.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Failed to delete: {e}")

                        # ---- EDIT MODE ----
                        if st.session_state.get(f"editing_config_{cid}", False):

                            if state_key not in st.session_state:
                                st.session_state[state_key] = [
                                    {
                                        'min_value': float(r['min_value']),
                                        'max_value': float(r['max_value']),
                                        'color': r['color']
                                    }
                                    for r in config.get('ranges', [])
                                ]

                            st.markdown(f"#### ✏️ Edit: {config.get('color_config_name')}")

                            edit_ranges = []

                            for idx, r in enumerate(st.session_state[state_key]):
                                ca, cb, cc, cd = st.columns([2, 2, 2, 1])

                                with ca:
                                    min_val = st.number_input(
                                        f"Min {idx+1}",
                                        value=float(r['min_value']),
                                        key=f"edit_min_{cid}_{idx}"
                                    )

                                with cb:
                                    max_val = st.number_input(
                                        f"Max {idx+1}",
                                        value=float(r['max_value']),
                                        key=f"edit_max_{cid}_{idx}"
                                    )

                                with cc:
                                    color_val = st.color_picker(
                                        f"Color {idx+1}",
                                        value=r['color'],
                                        key=f"edit_col_{cid}_{idx}"
                                    )

                                with cd:
                                    if st.button("❌", key=f"remove_{cid}_{idx}"):
                                        st.session_state[state_key].pop(idx)
                                        st.rerun()

                                edit_ranges.append({
                                    'min_value': Decimal(str(min_val)),
                                    'max_value': Decimal(str(max_val)),
                                    'color': color_val
                                })

                            col_a, col_b = st.columns(2)

                            with col_a:
                                if st.button("➕ Add Range", key=f"add_{cid}"):
                                    st.session_state[state_key].append({
                                        'min_value': 0.0,
                                        'max_value': 100.0,
                                        'color': '#ffffff'
                                    })
                                    st.rerun()

                            with col_b:
                                if st.button("Reset", key=f"reset_{cid}"):
                                    st.session_state[state_key] = [
                                        {
                                            'min_value': float(r['min_value']),
                                            'max_value': float(r['max_value']),
                                            'color': r['color']
                                        }
                                        for r in config.get('ranges', [])
                                    ]
                                    st.rerun()

                            edit_config_name = st.text_input("Configuration Name", value=config.get('color_config_name', ''), key=f"name_{cid}")
                            edit_locations = st.multiselect("Locations", ["tampere", "oulu", "rauma", "turku", "jyvaskyla", "vaasa", "seinajoki"], default=config.get('locations', []), key=f"loc_{cid}")
                            edit_dashboards = st.multiselect("Dashboards", ["price_dashboard", "historical_calendar"], default=config.get('dashboards', []), key=f"dash_{cid}")

                            if st.button("💾 Save Changes", key=f"save_{cid}"):
                                try:
                                    table_color.delete_item(Key={'color_config_name': config['color_config_name']})
                                    table_color.put_item(Item={
                                        'id': cid,
                                        'color_config_name': edit_config_name,
                                        'locations': edit_locations,
                                        'dashboards': edit_dashboards,
                                        'ranges': edit_ranges,
                                        'created_at': config.get('created_at', datetime.now().isoformat()),
                                        'created_by': config.get('created_by', 'admin')
                                    })
                                    st.session_state[f"editing_config_{cid}"] = False
                                    st.session_state.pop(state_key, None)
                                    st.success("Configuration updated!")
                                    time.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Failed to update: {e}")

                        st.divider()
                else:
                    st.info("No color configurations found")

            # ---- CREATE NEW CONFIG ----
            st.markdown("### Create New Color Configuration")

            default_color = [
                {'min_value': 0.0, 'max_value': 124.99, 'color': '#08306b'},
                {'min_value': 125.0, 'max_value': 134.99, 'color': '#2171b5'},
                {'min_value': 135.0, 'max_value': 144.99, 'color': '#a2cff8'},
                {'min_value': 145.0, 'max_value': 154.99, 'color': '#ffffff'},
                {'min_value': 155.0, 'max_value': 164.99, 'color': '#ffa0a0'},
                {'min_value': 165.0, 'max_value': 199.99, 'color': '#f86868'},
                {'min_value': 200.0, 'max_value': 249.99, 'color': '#d81919'},
                {'min_value': 250.0, 'max_value': 999999.0, 'color': '#000000'}
            ]

            if 'form_color_ranges' not in st.session_state:
                st.session_state.form_color_ranges = default_color

            color_config_name = st.text_input("Configuration Name", value="Default", key="config_name")
            selected_locations = st.multiselect("Locations", ["tampere", "oulu", "rauma", "turku", "jyvaskyla", "vaasa", "seinajoki"], default=["tampere"])
            selected_dashboards = st.multiselect("Dashboards", ["price_dashboard", "historical_calendar"], default=["price_dashboard"])

            st.markdown("#### Color Ranges")

            ranges_to_save = []
            for idx, r in enumerate(st.session_state.form_color_ranges):
                c1, c2, c3 = st.columns([2, 2, 2])
                with c1:
                    min_val = st.number_input(f"Min {idx}", value=float(r['min_value']), key=f"new_min_{idx}")
                with c2:
                    max_val = st.number_input(f"Max {idx}", value=float(r['max_value']), key=f"new_max_{idx}")
                with c3:
                    color_val = st.color_picker(f"Color {idx}", value=r['color'], key=f"new_col_{idx}")

                ranges_to_save.append({
                    'min_value': Decimal(str(min_val)),
                    'max_value': Decimal(str(max_val)),
                    'color': color_val
                })

            cadd, crem, creset = st.columns(3)

            with cadd:
                if st.button("➕ Add Range"):
                    st.session_state.form_color_ranges.append({'min_value': 0.0, 'max_value': 100.0, 'color': '#ffffff'})
                    st.rerun()

            with crem:
                if st.button("➖ Remove Last") and len(st.session_state.form_color_ranges) > 1:
                    st.session_state.form_color_ranges.pop()
                    st.rerun()

            with creset:
                if st.button("Reset"):
                    st.session_state.form_color_ranges = default_color
                    st.rerun()

            if st.button("💾 Save Configuration", type="primary"):
                try:
                    table_color.put_item(Item={
                        'id': str(uuid.uuid4()),
                        'color_config_name': color_config_name,
                        'locations': selected_locations,
                        'dashboards': selected_dashboards,
                        'ranges': ranges_to_save,
                        'created_at': datetime.now().isoformat(),
                        'created_by': st.session_state.get('authenticated_user', 'admin')
                    })
                    st.success("Saved!")
                    st.session_state.form_color_ranges = default_color
                    time.sleep(1)
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed: {e}")

        # ---- 4. Standardized Comparison Configuration ----
        with st.expander("📊 Standardized Comparison Configuration", expanded=False):
            st.caption("Average price top value - Price dashboard")

            if 'std_top_value' not in st.session_state:
                st.session_state.std_top_value = get_std_top_value()

            new_top_val = st.number_input(
                "Top value",
                value=int(st.session_state.std_top_value),
                min_value=50,
                step=50,
                key="std_top_number_input",
                label_visibility="collapsed"
            )
            st.session_state.std_top_value = int(new_top_val)

            if st.button("💾 Save for all users", key="std_save_btn", use_container_width=True, type="primary"):
                if save_std_top_value(st.session_state.std_top_value):
                    st.success(f"✅ Saved! All users will see max €{st.session_state.std_top_value}")
        
        # ---- 3. Zone Management ────────────────────────────────────────────
        with st.expander("🗺️ Zone Management", expanded=False):
            # ── Create New Zone ──────────────────────────────────────────────
            st.markdown("#### ➕ Create New Zone")
            with st.form("create_zone_form"):
                zf1, zf2 = st.columns([2, 2])
                with zf1: new_zone_name     = st.text_input("Zone Display Name", placeholder="e.g. Zone 1 City Center")
                with zf2: new_zone_location = st.selectbox("Location", ALL_LOCATIONS, key="new_zone_loc_sel")
                new_zone_hotels_raw = st.text_area("Hotels (one per line)", height=200,
                                                    placeholder="Scandic Tampere City\nSolo Sokos Hotel Torni Tampere\n...")
                create_zone_btn = st.form_submit_button("💾 Create Zone", type="primary", use_container_width=True)

            if create_zone_btn:
                if not new_zone_name.strip():
                    st.error("Zone name is required")
                else:
                    hotels_list = [h.strip() for h in new_zone_hotels_raw.splitlines() if h.strip()]
                    if save_zone(new_zone_name.strip(), new_zone_location, hotels_list):
                        st.success(f"✅ Zone '{new_zone_name}' created for {new_zone_location} with {len(hotels_list)} hotels!")
                        time.sleep(1); st.rerun()

            st.markdown("---")

            # ── Existing Zones ───────────────────────────────────────────────
            st.markdown("#### 📋 Existing Zones")
            filter_loc = st.selectbox("Filter by location", ["All"] + ALL_LOCATIONS, key="zone_admin_filter_loc")
            all_zones     = get_all_zones()
            display_zones = [z for z in all_zones if filter_loc == "All" or z.get('location') == filter_loc]

            if not display_zones:
                st.info("No zones found. Create one above.")
            else:
                for zone in display_zones:
                    z_name = zone['zone_name']
                    z_loc  = zone['location']
                    # State keys use BOTH parts of the composite PK to guarantee uniqueness
                    sk = f"{z_name}__{z_loc}"

                    lc, ec, dc = st.columns([5, 1, 1])
                    with lc:
                        st.markdown(f"**{z_name}** · `{z_loc}` · {len(zone.get('hotels',[]))} hotels")
                    with ec:
                        if st.button("✏️ Edit", key=f"zone_edit_{sk}", use_container_width=True):
                            st.session_state[f"zone_editing_{sk}"] = not st.session_state.get(f"zone_editing_{sk}", False)
                    with dc:
                        if st.button("🗑️", key=f"zone_del_{sk}", use_container_width=True):
                            st.session_state[f"zone_del_confirm_{sk}"] = True

                    # Delete confirmation
                    if st.session_state.get(f"zone_del_confirm_{sk}", False):
                        st.warning(f"⚠️ Delete **{z_name}** (`{z_loc}`)?")
                        yd, nd, _ = st.columns([1, 1, 3])
                        with yd:
                            if st.button("✅ Yes", key=f"zone_del_yes_{sk}", use_container_width=True):
                                if delete_zone(z_name, z_loc):
                                    st.success("Zone deleted")
                                    st.session_state[f"zone_del_confirm_{sk}"] = False
                                    time.sleep(1); st.rerun()
                        with nd:
                            if st.button("❌ Cancel", key=f"zone_del_no_{sk}", use_container_width=True):
                                st.session_state[f"zone_del_confirm_{sk}"] = False; st.rerun()

                    # Inline edit form
                    if st.session_state.get(f"zone_editing_{sk}", False):
                        with st.form(f"edit_zone_form_{sk}"):
                            st.markdown(f"##### ✏️ Editing: **{z_name}** (`{z_loc}`)")
                            st.caption(
                                "⚠️ Changing Zone Name or Location changes the PK (`zone_name#location`). "
                                "The old record will be deleted and a new one created automatically."
                            )
                            ef1, ef2 = st.columns([2, 2])
                            with ef1: edit_z_name  = st.text_input("Zone Name",   value=z_name, key=f"ez_name_{sk}")
                            with ef2: edit_z_loc   = st.selectbox("Location", ALL_LOCATIONS,
                                                                    index=ALL_LOCATIONS.index(z_loc) if z_loc in ALL_LOCATIONS else 0,
                                                                    key=f"ez_loc_{sk}")
                            edit_hotels_raw = st.text_area("Hotels (one per line)",
                                                            value="\n".join(zone.get('hotels',[])),
                                                            height=250, key=f"ez_hotels_{sk}")
                            save_edit_btn = st.form_submit_button("💾 Save Changes", type="primary", use_container_width=True)

                        if save_edit_btn:
                            edited_hotels = [h.strip() for h in edit_hotels_raw.splitlines() if h.strip()]
                            key_changed   = (edit_z_name.strip() != z_name) or (edit_z_loc != z_loc)
                            if key_changed:
                                delete_zone(z_name, z_loc)   # remove old composite key first
                            if save_zone(edit_z_name.strip(), edit_z_loc, edited_hotels):
                                st.success("✅ Zone updated!")
                                st.session_state[f"zone_editing_{sk}"] = False
                                time.sleep(1); st.rerun()

                    st.divider()

        with st.expander("📧 Email Management", expanded=False):
            st.caption(
                "Global email list — addresses saved here appear as selectable "
                "recipients in the Matrix Automation tab."
            )
 
            # Add
            st.markdown("#### ➕ Add Email Address")
            em_c1, em_c2 = st.columns([4, 1])
            with em_c1:
                new_email_input = st.text_input(
                    "Email address",
                    placeholder="name@company.com",
                    label_visibility="collapsed",
                    key="new_email_input"
                )
            with em_c2:
                add_email_btn = st.button(
                    "Add", type="primary",
                    use_container_width=True,
                    key="add_email_btn"
                )
 
            if add_email_btn:
                val = new_email_input.strip().lower()
                if not val or "@" not in val:
                    st.error("Enter a valid email address.")
                else:
                    if save_email(val):
                        st.success(f"✅ **{val}** added.")
                        time.sleep(0.5)
                        st.rerun()
 
            st.markdown("---")
 
            # List
            st.markdown("#### 📋 Saved Email Addresses")
            current_emails = get_saved_emails()
 
            if not current_emails:
                st.info("No email addresses saved yet.")
            else:
                st.caption(f"{len(current_emails)} address(es) in the global list.")
                for em in current_emails:
                    ec1, ec2 = st.columns([5, 1])
                    with ec1:
                        st.markdown(f"📧 `{em}`")
                    with ec2:
                        if st.button("🗑️", key=f"del_email_{em}",
                                     use_container_width=True):
                            if delete_email(em):
                                st.success(f"Removed **{em}**.")
                                time.sleep(0.5)
                                st.rerun()